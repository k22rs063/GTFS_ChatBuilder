"""時刻表ファイル (CSV / xlsx) の入力ヘルパ。

tools/stop_times.py と app.py (確認層プレビュー) の両方から利用される。
1 ファイルから (シート名, CSV テキスト) のリストを取り出す共通インタフェース。

責務は読み込みのみ。テンプレ形式への変換は timetable_normalizer 側、
GTFS 出力は stop_times 側の責務。
"""

from __future__ import annotations

import csv
from io import StringIO
from pathlib import Path

import openpyxl

from gtfs_chatbuilder.processors.encoding import read_text_auto


def read_timetable_sources(path: Path) -> list[tuple[str, str]]:
    """ファイルパスから (シート名, CSV テキスト) のリストを返す。

    - CSV: 1 要素のリスト。utf-8-sig / cp932 を順に試行
    - xlsx / xls: 全シートのリスト (空シートは除外)

    Raises:
        ValueError: 未対応の拡張子。
    """
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        return _xlsx_sources(path)
    if suffix == ".csv":
        return _csv_sources(path)
    raise ValueError(
        f"未対応の拡張子です ({suffix})。"
        ".csv / .xlsx / .xls のいずれかを指定してください。"
    )


def _csv_sources(path: Path) -> list[tuple[str, str]]:
    return [(path.name, read_text_auto(path))]


def _xlsx_sources(path: Path) -> list[tuple[str, str]]:
    """xlsx の全シートを (sheet_name, csv_text) のリストにする。

    空セルは "" に、日時セルは str() に変換。空白だけのシートは除外。
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        result: list[tuple[str, str]] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            buf = StringIO()
            writer = csv.writer(buf, lineterminator="\n")
            had_content = False
            for row in ws.iter_rows(values_only=True):
                cells: list[str] = []
                for v in row:
                    if v is None:
                        cells.append("")
                    elif isinstance(v, str):
                        cells.append(v)
                        if v.strip():
                            had_content = True
                    else:
                        cells.append(str(v))
                        had_content = True
                writer.writerow(cells)
            text = buf.getvalue()
            if text.strip() and had_content:
                result.append((sheet_name, text))
        return result
    finally:
        wb.close()
