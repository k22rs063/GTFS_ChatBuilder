"""区役所担当者に配布する入力用 Excel テンプレートを動的生成する。

設計指針 ([[Excel入力と座標の設計指針]]):
- 用途別にファイルを分割
- 日本語ヘッダー
- Streamlit の st.download_button で配信 (常に最新版が手に入る)

各 build_*_template() は .xlsx のバイト列を返す。
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

# stops テンプレートの列構成 (日本語ヘッダー, 補足コメント)
_STOPS_COLUMNS: list[tuple[str, str]] = [
    ("停留所名", "必須。例: 市役所前"),
    ("停留所ID", "任意。空欄なら自動採番 (stop_1, stop_2...)"),
    ("読み仮名", "任意。例: しやくしょまえ"),
    ("緯度", "任意。例: 33.611111 (空欄ならKMLから補完)"),
    ("経度", "任意。例: 131.011111 (空欄ならKMLから補完)"),
    ("乗り場番号", "任意。例: 1"),
    ("運賃エリアID", "任意。ゾーン制/対キロ運賃の場合に設定"),
]


def build_stops_template() -> bytes:
    """停留所マスタ入力用テンプレート (stops.xlsx) のバイト列を返す。

    1行目: 日本語ヘッダー
    2行目: 記入例 (グレー想定、担当者が消して使う)
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "停留所マスタ"

    headers = [col[0] for col in _STOPS_COLUMNS]
    sheet.append(headers)

    # 記入例の行
    example = [
        "市役所前",
        "",
        "しやくしょまえ",
        "33.611111",
        "131.011111",
        "1",
        "",
    ]
    sheet.append(example)

    # 列幅を見やすく調整
    for idx, _ in enumerate(headers, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=idx).column_letter].width = 16

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# routes テンプレートの列構成
_ROUTES_COLUMNS: list[tuple[str, str]] = [
    ("路線名", "必須。例: 市役所線"),
    ("系統番号", "任意。車両に表示される系統番号。例: 東16"),
    ("路線ID", "任意。空欄なら自動採番 (route_1, route_2...)"),
    ("種別", "任意。バス/鉄道/路面電車/船 など。空欄ならバス"),
]


def build_routes_template() -> bytes:
    """路線マスタ入力用テンプレート (routes.xlsx) のバイト列を返す。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "路線マスタ"

    headers = [col[0] for col in _ROUTES_COLUMNS]
    sheet.append(headers)
    sheet.append(["市役所線", "", "", "バス"])

    for idx, _ in enumerate(headers, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=idx).column_letter].width = 16

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
