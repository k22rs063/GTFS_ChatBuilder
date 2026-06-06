"""停留所マスタ関連ツール。

- import_stops_from_excel: 日本語ヘッダーExcel → stops.txt 生成
- fill_stop_coordinates_from_kml: 座標が空の停留所を KML のマーカーから補完

GTFS-JP v4 必須ファイル stops.txt を扱う。

設計指針 ([[Excel入力と座標の設計指針]]):
- 日本語ヘッダーのテンプレート、内部で GTFSフィールド名にマッピング
- 座標 (緯度/経度) が空でも取り込みは許可、警告を出す (後で KML補完)
- Phase 1 は location_type=0 (乗り場) のみ。親station等の階層は Phase 2

KML は Google マイマップで停留所をマーカー(Point)で打ったものを想定。
マーカー名 = 停留所名。経路ライン(LineString)が同じ KML に混在してもよい
(座標補完では Point のみ拾う)。

LLM はファイル名を引数に詰めるだけ。値は Excel/KML から決定論的に読む
([[LLM is router only]])。
"""

from __future__ import annotations

import xml.etree.ElementTree as ET

from openpyxl import load_workbook

from langchain.tools import tool

from gtfs_chatbuilder.gtfs_writer import write_gtfs_csv
from gtfs_chatbuilder.paths import WORKSPACE_DIR
from gtfs_chatbuilder.processors.encoding import read_text_auto
from gtfs_chatbuilder.processors.stops_normalizer import normalize_stops
from gtfs_chatbuilder.progress import load_progress, save_progress
from gtfs_chatbuilder.validators.base import read_csv_records


@tool
def generate_stops_from_csv(input_csv_filename: str) -> str:
    """自治体が持っている停留所情報 CSV から GTFS の stops.txt を生成する。

    自治体の停留所一覧 CSV (バス停名 + 緯度経度 等を含むもの) を読み込み、
    GTFS の stops.txt (stop_id, stop_name, stop_lat, stop_lon) に変換する。

    対応する入力:
    - エンコード: cp932 / utf-8 / utf-8-sig (BOM 有無問わず) を自動判別
    - ヘッダ別名: 「停留所名/バス停名/停留所」「緯度経度/緯度/経度」 等
    - 緯度経度: 十進数 (33.6158) も DMS (33°36'56.9"N 131°03'52.5"E) も両対応
    - stop_id: 明示列 > No. 列 > 自動連番 (S1, S2, ...) の順で決まる

    出力先は workspace/stops.txt 固定。
    LLM はファイル名を引数に詰めるだけで、stop_id や座標は決定論的に作る。

    Args:
        input_csv_filename: 停留所情報 CSV のファイル名 (例: "停留所情報.csv")

    Returns:
        生成結果のサマリ (件数、座標未設定数など)
    """
    input_path = WORKSPACE_DIR / input_csv_filename
    if not input_path.exists():
        return f"エラー: ファイルが見つかりません: {input_path}"

    try:
        text = read_text_auto(input_path)
    except UnicodeDecodeError as e:
        return f"エラー: 停留所 CSV のエンコード判定に失敗: {e}"

    try:
        normalized = normalize_stops(text)
    except ValueError as e:
        return f"エラー: 停留所 CSV の変換に失敗: {e}"

    output = WORKSPACE_DIR / "stops.txt"
    output.write_text(normalized, encoding="utf-8")

    lines = [r for r in normalized.split("\n") if r]
    count = len(lines) - 1  # header を除く
    # 座標未設定をカウント
    missing_coords = 0
    for ln in lines[1:]:
        cells = ln.split(",")
        if len(cells) >= 4 and (not cells[2].strip() or not cells[3].strip()):
            missing_coords += 1

    summary = (
        f"stops.txt を生成しました ({count}件の停留所)。\n"
        f"出力先: {output}"
    )
    if missing_coords:
        summary += (
            f"\n⚠ {missing_coords}件の停留所で座標が未設定です。"
            "元データに緯度経度を追加するか、KML から補完してください。"
        )
    return summary

# 日本語ヘッダー → GTFSフィールド名 のマッピング。
# キーは正規化 (前後空白除去・小文字化) して比較する。
_HEADER_MAP = {
    "停留所名": "stop_name",
    "停留所id": "stop_id",
    "緯度": "stop_lat",
    "経度": "stop_lon",
    "乗り場番号": "platform_code",
    "運賃エリアid": "zone_id",
}
# 「読み仮名」列は translations ツールで扱うため、ここでは取り込まない (無視)。

# stops.txt の出力フィールド順 (Phase 1 スコープ)
_OUTPUT_HEADER = [
    "stop_id",
    "stop_name",
    "stop_lat",
    "stop_lon",
    "zone_id",
    "location_type",
    "platform_code",
]


@tool
def import_stops_from_excel(excel_filename: str) -> str:
    """停留所マスタ Excel を読み込んで stops.txt を生成する (GTFS-JP v4 必須ファイル)。

    Excel は workspace フォルダ内に置かれている前提。日本語ヘッダーのテンプレート
    (停留所名/停留所ID/緯度/経度/乗り場番号/運賃エリアID) を想定する。
    「停留所名」列は必須。緯度・経度が空の停留所があっても取り込みは行い、
    その件数を警告として返す (後で KML から座標を補完できる)。

    Args:
        excel_filename: 停留所マスタExcelのファイル名 (例: "stops.xlsx")

    Returns:
        生成結果のサマリ (取り込み件数、座標未設定の警告など)
    """
    excel_path = WORKSPACE_DIR / excel_filename
    if not excel_path.exists():
        return (
            f"エラー: Excelファイルが見つかりません: {excel_path}\n"
            "workspace フォルダにファイルを置いてください。"
        )

    try:
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 - openpyxl の各種例外を網羅的に拾う
        return f"エラー: Excelファイルを開けませんでした: {e}"

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        return "エラー: Excelが空です。"

    header_cells = rows[0]
    col_index = _map_header(header_cells)

    if "stop_name" not in col_index:
        return (
            "エラー: 「停留所名」列が見つかりません。"
            "テンプレートの1行目に「停留所名」列を用意してください。"
        )

    data_rows = rows[1:]
    output_rows: list[list[str]] = []
    auto_id_counter = 0
    missing_coord_names: list[str] = []
    skipped_blank = 0

    for raw in data_rows:
        stop_name = _cell(raw, col_index.get("stop_name"))
        if not stop_name:
            skipped_blank += 1
            continue

        stop_id = _cell(raw, col_index.get("stop_id"))
        if not stop_id:
            auto_id_counter += 1
            stop_id = f"stop_{auto_id_counter}"

        stop_lat = _cell(raw, col_index.get("stop_lat"))
        stop_lon = _cell(raw, col_index.get("stop_lon"))
        if not stop_lat or not stop_lon:
            missing_coord_names.append(stop_name)

        output_rows.append(
            [
                stop_id,
                stop_name,
                stop_lat,
                stop_lon,
                _cell(raw, col_index.get("zone_id")),
                "0",  # location_type: Phase 1 は乗り場固定
                _cell(raw, col_index.get("platform_code")),
            ]
        )

    if not output_rows:
        return "エラー: 有効な停留所データがありませんでした (停留所名がすべて空です)。"

    output = WORKSPACE_DIR / "stops.txt"
    write_gtfs_csv(output, _OUTPUT_HEADER, output_rows)

    _update_progress(excel_filename, len(output_rows), missing_coord_names)

    lines = [
        f"stops.txt を生成しました ({len(output_rows)}件の停留所)。",
        f"出力先: {output}",
    ]
    if skipped_blank:
        lines.append(f"※ 停留所名が空の {skipped_blank} 行はスキップしました。")
    if missing_coord_names:
        preview = "、".join(missing_coord_names[:5])
        more = "" if len(missing_coord_names) <= 5 else " 他"
        lines.append(
            f"⚠ 座標 (緯度/経度) が未設定の停留所が {len(missing_coord_names)} 件あります: "
            f"{preview}{more}。"
            "KMLファイルから座標を補完するか、Excelに座標を追記してください。"
        )
    return "\n".join(lines)


def _map_header(header_cells: tuple) -> dict[str, int]:
    """ヘッダー行を走査し、GTFSフィールド名 → 列インデックス の辞書を返す。"""
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_cells):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        field = _HEADER_MAP.get(key)
        if field and field not in mapping:
            mapping[field] = idx
    return mapping


def _cell(row: tuple, index: int | None) -> str:
    """行から指定列の値を取り出し、文字列化して前後空白を除去する。"""
    if index is None or index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    return str(value).strip()


def _update_progress(
    excel_filename: str, stop_count: int, missing_coord_names: list[str]
) -> None:
    """メタファイルの stops ステップを更新する。"""
    progress = load_progress()
    step = progress.steps.get("stops")
    if step is None:
        return
    step.source_files = [excel_filename]
    step.fields_set = ["stop_id", "stop_name", "location_type"]
    if missing_coord_names:
        # 座標未設定は in_progress 扱い (zip化前に解消が必要)
        step.status = "in_progress"
        step.fields_missing = [
            f"{len(missing_coord_names)}件の停留所で座標(stop_lat/stop_lon)が未設定"
        ]
    else:
        step.status = "completed"
        step.fields_missing = []
    save_progress(progress)


# --- 座標補完 (KML マーカー) ---------------------------------------------


@tool
def fill_stop_coordinates_from_kml(kml_filename: str) -> str:
    """KMLファイルのマーカーから、stops.txt の空の座標を補完する。

    Google マイマップで停留所をマーカー(ピン)で打ち、ピン名を停留所名に
    した KML を想定する。停留所名が完全一致するマーカーの座標を、
    stops.txt の座標が空の行に書き込む。経路ライン(LineString)が同じ KML に
    含まれていても無視する。

    事前に import_stops_from_excel で stops.txt を作っておく必要がある。

    Args:
        kml_filename: 停留所マーカーを含む KML のファイル名 (例: "stops.kml")

    Returns:
        補完結果のサマリ (補完した件数、まだ空の停留所、未使用のマーカー名)
    """
    stops_path = WORKSPACE_DIR / "stops.txt"
    if not stops_path.exists():
        return (
            "エラー: stops.txt がまだありません。"
            "先に import_stops_from_excel で停留所マスタを取り込んでください。"
        )

    kml_path = WORKSPACE_DIR / kml_filename
    if not kml_path.exists():
        return (
            f"エラー: KMLファイルが見つかりません: {kml_path}\n"
            "workspace フォルダにファイルを置いてください。"
        )

    try:
        kml_points = _parse_kml_points(kml_path)
    except ET.ParseError as e:
        return f"エラー: KMLの解析に失敗しました: {e}"

    if not kml_points:
        return (
            "KML 内に停留所マーカー(Point)が見つかりませんでした。"
            "Google マイマップで停留所をマーカーとして打ち、名前を付けてください。"
        )

    records = read_csv_records(stops_path)
    if not records:
        return "エラー: stops.txt にデータがありません。"

    filled: list[str] = []
    still_missing: list[str] = []
    used_marker_names: set[str] = set()

    for rec in records:
        name = (rec.get("stop_name") or "").strip()
        lat = (rec.get("stop_lat") or "").strip()
        lon = (rec.get("stop_lon") or "").strip()
        if lat and lon:
            continue  # 既に座標あり
        match = kml_points.get(name)
        if match is not None:
            rec["stop_lat"], rec["stop_lon"] = match
            filled.append(name)
            used_marker_names.add(name)
        else:
            still_missing.append(name)

    if not filled:
        detail = (
            "停留所名が KML のマーカー名と完全一致しませんでした。"
            "名前の表記 (空白・括弧など) が一致しているか確認してください。"
        )
        return f"座標を補完できる停留所がありませんでした。\n{detail}"

    # _OUTPUT_HEADER の順で stops.txt を書き直す
    output_rows = [[rec.get(col, "") for col in _OUTPUT_HEADER] for rec in records]
    write_gtfs_csv(stops_path, _OUTPUT_HEADER, output_rows)

    _refresh_stops_progress(kml_filename, still_missing)

    lines = [
        f"KML から {len(filled)} 件の停留所の座標を補完しました: "
        + "、".join(filled[:8])
        + ("" if len(filled) <= 8 else " 他"),
    ]
    if still_missing:
        preview = "、".join(still_missing[:5])
        more = "" if len(still_missing) <= 5 else " 他"
        lines.append(
            f"⚠ まだ座標が未設定の停留所が {len(still_missing)}件あります: "
            f"{preview}{more}"
        )
    unused = sorted(set(kml_points) - used_marker_names)
    if unused:
        preview = "、".join(unused[:5])
        more = "" if len(unused) <= 5 else " 他"
        lines.append(
            f"(参考) stops.txt の停留所と一致しなかった KML マーカー: "
            f"{preview}{more}"
        )
    return "\n".join(lines)


def _parse_kml_points(kml_path) -> dict[str, tuple[str, str]]:
    """KML を解析し、マーカー名 → (緯度, 経度) の辞書を返す。

    Point を持つ Placemark のみ対象。LineString 等は無視する。
    名前空間の有無に依存しないよう、タグのローカル名で判定する。
    """
    tree = ET.parse(kml_path)
    root = tree.getroot()
    points: dict[str, tuple[str, str]] = {}

    for placemark in root.iter():
        if _localname(placemark.tag) != "Placemark":
            continue
        name = ""
        coord_text = ""
        for child in placemark.iter():
            local = _localname(child.tag)
            if local == "name" and not name:
                name = (child.text or "").strip()
            elif local == "coordinates" and not coord_text:
                # Point 配下の coordinates のみ採用したいので、
                # 親が Point かを確認するため後段でフィルタする
                coord_text = (child.text or "").strip()
        # Point を持つか確認 (LineString の coordinates を拾わないため)
        has_point = any(
            _localname(el.tag) == "Point" for el in placemark.iter()
        )
        if not name or not has_point:
            continue
        latlon = _coord_to_latlon(coord_text)
        if latlon and name not in points:
            points[name] = latlon
    return points


def _localname(tag: str) -> str:
    """名前空間付きタグから要素名のみを取り出す ({ns}Placemark -> Placemark)。"""
    return tag.rsplit("}", 1)[-1]


def _coord_to_latlon(coord_text: str) -> tuple[str, str] | None:
    """KML の coordinates ("経度,緯度,高度") を (緯度, 経度) に変換する。"""
    if not coord_text:
        return None
    # Point の coordinates は通常1点だが、改行や空白が混じる場合に備え先頭を採る
    first = coord_text.split()[0] if coord_text.split() else coord_text
    parts = first.split(",")
    if len(parts) < 2:
        return None
    lon, lat = parts[0].strip(), parts[1].strip()
    if not lon or not lat:
        return None
    return (lat, lon)


def _refresh_stops_progress(kml_filename: str, still_missing: list[str]) -> None:
    """座標補完後、stops ステップの進捗を更新する。"""
    progress = load_progress()
    step = progress.steps.get("stops")
    if step is None:
        return
    if kml_filename not in step.source_files:
        step.source_files = [*step.source_files, kml_filename]
    if still_missing:
        step.status = "in_progress"
        step.fields_missing = [
            f"{len(still_missing)}件の停留所で座標(stop_lat/stop_lon)が未設定"
        ]
    else:
        step.status = "completed"
        step.fields_missing = []
    save_progress(progress)
