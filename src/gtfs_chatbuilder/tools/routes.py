"""路線マスタ Excel 取り込みツール (import_routes_from_excel)。

GTFS-JP v4 必須ファイル routes.txt を、日本語ヘッダーの Excel から生成する。

設計指針 ([[Excel入力と座標の設計指針]]): stops と同じ取り込みパターン。
- agency_id は agency.txt から自動で引く (現状は 1 workspace = 1 事業者)
- route_type は日本語 (「バス」等) を数値にマッピング。空ならバス=3
- route_short_name / route_long_name のどちらか必須 (v4 仕様)

LLM はファイル名を引数に詰めるだけ ([[LLM is router only]])。
"""

from __future__ import annotations

from openpyxl import load_workbook

from langchain.tools import tool

from gtfs_chatbuilder.gtfs_writer import write_gtfs_csv
from gtfs_chatbuilder.paths import WORKSPACE_DIR
from gtfs_chatbuilder.progress import load_progress, save_progress
from gtfs_chatbuilder.validators.base import read_csv_records

# 日本語ヘッダー → GTFSフィールド名 (正規化キーで比較)
_HEADER_MAP = {
    "路線名": "route_long_name",
    "系統番号": "route_short_name",
    "路線id": "route_id",
    "種別": "_route_type_jp",
}

# 種別の日本語 → route_type 数値 (v4 値の設定方法)
_ROUTE_TYPE_MAP = {
    "路面電車": "0",
    "ライトレール": "0",
    "地下鉄": "1",
    "鉄道": "2",
    "電車": "2",
    "バス": "3",
    "コミュニティバス": "3",
    "デマンド": "3",
    "フェリー": "4",
    "船": "4",
    "旅客船": "4",
}
_DEFAULT_ROUTE_TYPE = "3"  # バス

_OUTPUT_HEADER = [
    "route_id",
    "agency_id",
    "route_short_name",
    "route_long_name",
    "route_type",
]


@tool
def import_routes_from_excel(excel_filename: str) -> str:
    """路線マスタ Excel を読み込んで routes.txt を生成する (GTFS-JP v4 必須ファイル)。

    Excel は workspace フォルダ内に置かれている前提。日本語ヘッダー
    (路線名/系統番号/路線ID/種別) を想定する。「路線名」列は必須。
    agency_id は agency.txt から自動取得する。

    Args:
        excel_filename: 路線マスタExcelのファイル名 (例: "routes.xlsx")

    Returns:
        生成結果のサマリ。
    """
    excel_path = WORKSPACE_DIR / excel_filename
    if not excel_path.exists():
        return (
            f"エラー: Excelファイルが見つかりません: {excel_path}\n"
            "workspace フォルダにファイルを置いてください。"
        )

    agency_id, agency_warning = _resolve_agency_id()

    try:
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        return f"エラー: Excelファイルを開けませんでした: {e}"

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        return "エラー: Excelが空です。"

    col_index = _map_header(rows[0])
    if "route_long_name" not in col_index and "route_short_name" not in col_index:
        return (
            "エラー: 「路線名」列が見つかりません。"
            "テンプレートの1行目に「路線名」列を用意してください。"
        )

    output_rows: list[list[str]] = []
    auto_id = 0
    skipped = 0
    nameless = 0

    for raw in rows[1:]:
        long_name = _cell(raw, col_index.get("route_long_name"))
        short_name = _cell(raw, col_index.get("route_short_name"))
        if not long_name and not short_name:
            skipped += 1
            continue

        route_id = _cell(raw, col_index.get("route_id"))
        if not route_id:
            auto_id += 1
            route_id = f"route_{auto_id}"

        route_type_jp = _cell(raw, col_index.get("_route_type_jp"))
        route_type = _resolve_route_type(route_type_jp)

        output_rows.append(
            [route_id, agency_id, short_name, long_name, route_type]
        )

    if not output_rows:
        return "エラー: 有効な路線データがありませんでした (路線名がすべて空です)。"

    output = WORKSPACE_DIR / "routes.txt"
    write_gtfs_csv(output, _OUTPUT_HEADER, output_rows)

    _update_progress(excel_filename, agency_id)

    lines = [
        f"routes.txt を生成しました ({len(output_rows)}件の路線)。",
        f"出力先: {output}",
    ]
    if agency_warning:
        lines.append(agency_warning)
    if skipped:
        lines.append(f"※ 路線名が空の {skipped} 行はスキップしました。")
    return "\n".join(lines)


def _resolve_agency_id() -> tuple[str, str]:
    """agency.txt から agency_id を取得する。

    Returns:
        (agency_id, 警告文字列)。agency.txt が無ければプレースホルダと警告を返す。
    """
    agency_path = WORKSPACE_DIR / "agency.txt"
    if not agency_path.exists():
        return (
            "agency_1",
            "⚠ agency.txt がまだありません。agency_id を仮の値 'agency_1' に "
            "しました。先に事業者情報 (set_agency) を登録することを推奨します。",
        )
    records = read_csv_records(agency_path)
    if not records:
        return ("agency_1", "⚠ agency.txt が空です。agency_id を 'agency_1' にしました。")
    agency_id = (records[0].get("agency_id") or "").strip()
    if not agency_id:
        return ("agency_1", "⚠ agency.txt に agency_id がありません。'agency_1' にしました。")
    return (agency_id, "")


def _resolve_route_type(route_type_jp: str) -> str:
    """種別の日本語表記を route_type 数値に変換する。空・不明ならバス(3)。"""
    value = route_type_jp.strip()
    if not value:
        return _DEFAULT_ROUTE_TYPE
    # 既に数値ならそのまま採用
    if value.isdigit():
        return value
    return _ROUTE_TYPE_MAP.get(value, _DEFAULT_ROUTE_TYPE)


def _map_header(header_cells: tuple) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_cells):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        field = _HEADER_MAP.get(key)
        if field and field not in mapping:
            mapping[field] = idx
    return mapping


def _cell(row: tuple, index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    return str(value).strip()


def _update_progress(excel_filename: str, agency_id: str) -> None:
    """メタファイルの routes ステップを更新する。"""
    progress = load_progress()
    step = progress.steps.get("routes")
    if step is None:
        return
    step.status = "completed"
    step.source_files = [excel_filename]
    step.fields_set = ["route_id", "agency_id", "route_type"]
    step.fields_missing = []
    save_progress(progress)
